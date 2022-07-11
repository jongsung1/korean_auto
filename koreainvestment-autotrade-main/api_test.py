from numpy import diff
import pandas as pd
import mojito
import all_companys_list as COM ####### 종목 코드 리스트
import requests
import json
import yaml
import config as c
from tqdm import tqdm ####### 진행률 표시 - for문에 사용
import openpyxl

#from search_company import PRICE_TEMP ####### 엑셀 write에 사용


with open('config.yaml', encoding='UTF-8') as f:
    _cfg = yaml.load(f, Loader=yaml.FullLoader)
APP_KEY = _cfg['APP_KEY']
APP_SECRET = _cfg['APP_SECRET']
ACCESS_TOKEN = ""
CANO = _cfg['CANO']
ACNT_PRDT_CD = _cfg['ACNT_PRDT_CD']
DISCORD_WEBHOOK_URL = _cfg['DISCORD_WEBHOOK_URL']
URL_BASE = _cfg['URL_BASE']
percent = _cfg['PERCENT']   ##현재가가 최고가의 $percent% 범위 이내(0 ~ 1)
symbol_list = _cfg['SYMBLO_LIST']   # 매수 희망 종목 리스트
k = _cfg['K']   ## 변동성 상수

def get_access_token():
    """토큰 발급"""
    headers = {"content-type":"application/json"}
    body = {"grant_type":"client_credentials",
    "appkey":APP_KEY, 
    "appsecret":APP_SECRET}
    PATH = "oauth2/tokenP"
    URL = f"{URL_BASE}/{PATH}"
    res = requests.post(URL, headers=headers, data=json.dumps(body))
    ACCESS_TOKEN = res.json()["access_token"]
    return ACCESS_TOKEN
    
def get_current_price(code="005930",symbol="aspr_unit"):
    """호가 조회"""
    PATH = "/uapi/domestic-stock/v1/quotations/inquire-price"
    URL = f"{URL_BASE}/{PATH}"
    headers = {"Content-Type":"application/json", 
            "authorization": f"Bearer {ACCESS_TOKEN}",
            "appkey":APP_KEY,
            "appsecret":APP_SECRET,
            "tr_id":"FHKST01010100"}
    params = {
    "fid_cond_mrkt_div_code":"J",
    "fid_input_iscd":code,
    }
    res = requests.get(URL, headers=headers, params=params)
    #return int(res.json()['output'][symbol])
    #return res.json()['output'][symbol]
    try:
        return res.json()['output'][symbol]
    except KeyError:
        return 0

def exit_typeError(val_name,val):
    """에러 try except"""
    try:
        val_name=val
    except TypeError:
        val_name=-1
    

ACCESS_TOKEN = get_access_token()
UNIT_PRICE=get_current_price("000660","aspr_unit")  ## 종목의 호가
EPS=get_current_price("000660","eps")   ## 종목의 EPS
BPS=get_current_price("000660","bps")   ## 종목의 BPS
PER=get_current_price("000660","per")   ## 종목의 EPS
PBR=get_current_price("000660","pbr")   ## 종목의 PBR
MARKET_CAP=get_current_price("000660","hts_avls")   ## 종목의 시가총액

####### 엑셀 저장을 위한 파트
row=1
wb = openpyxl.Workbook()
ws = wb.active

####### 엑셀 저장 함수
def write_excel(company_name,MARKET_CAP,price,hgpr1,differnce,row):
    ws.cell(row=row+1, column=1).value = company_name               ## 회사명
    ws.cell(row=row+1, column=2).value = MARKET_CAP                 ## 시가총액
    ws.cell(row=row+1, column=3).value = price                      ## 현재가
    ws.cell(row=row+1, column=4).value = hgpr1                      ## 최고가
    ws.cell(row=row+1, column=5).value = differnce                  ## 가격차
    
####### 엑셀 헤더
write_excel("회사명","시가총액","현재가","최고가","가격차",0)

### 가운데자리 검색
broker = mojito.KoreaInvestment(api_key=APP_KEY, api_secret=APP_SECRET)
for company in tqdm(COM.tickers.keys()):
    UNIT_PRICE=get_current_price(company,"aspr_unit")   ## 호가
    diff_val = int(2*UNIT_PRICE)

    resp = broker.fetch_daily_price(company)
    #price_t = resp['output'][0]['stck_clpr']  ## 현재가
    PRICE_TEMP=get_current_price(company,"stck_prpr")    ## 현재가
    price=int(PRICE_TEMP)
    company_name=COM.tickers[company]
    j = len(resp['output']) ## 총 자료의 길이

    ### 종목의 고가를 hprice 리스트에 저장
    hprice = []
    for i in range(0, j):
        hprice.append(resp['output'][i]['stck_hgpr'])

    hprice.sort()   ## hprice 크기 순 정렬
    hgpr1 = hprice[-1]    ## 최고가
    hgpr2 = hprice[-2]    ## 2번째 최고가

    target = int(hgpr1) * float(percent)

    differnce = int(hgpr1) - int(hgpr2)
    
    CHECK_VAL1 = 0
    CHECK_VAL2 = 0
    #exit_typeError(CHECK_VAL1,differnce-diff_val)
    #exit_typeError(CHECK_VAL2,hgpr1-price)
    try:
        CHECK_VAL1=differnce-diff_val
    except TypeError:
        CHECK_VAL1=-1

    try:
        CHECK_VAL2=hgpr1-price
    except TypeError:
        CHECK_VAL2=-1

    if CHECK_VAL1 < 0:
        if CHECK_VAL2 > 0:
            if (float(target)-int(price) < 0):  ##현재가가 최고가의 일정 범위 이내
                print("%s 종목 현재가 : %s , 최고가 : %s , 최고값 간 차이 : %s" % (company_name,price,hgpr1,differnce))
                ####### 엑셀 입력 함수 호출
                write_excel(company_name,MARKET_CAP,price,hgpr1,differnce,row)
                row = row + 1
    
wb.save(f"./center_value.xlsx")




# df = pd.DataFrame(resp['output'])
# dt = pd.to_datetime(df['stck_bsop_date'], format="%Y%m%d")
# df.set_index(dt, inplace=True)
# df = df[['stck_oprc', 'stck_hgpr', 'stck_lwpr', 'stck_clpr']]
# df.columns = ['open', 'high', 'low', 'close']
# df.index.name = "date"
#print(df)


# print(resp['output'][0]['stck_bsop_date']) ## 특정 종목정보 날짜 0 1 2 3 4 ...
# print(resp['output'][0]['stck_oprc']) ## 특정 종목의 시가
# print(resp['output'][0]['stck_hgpr']) ## 특정 종목의 고가
# print(resp['output'][0]['stck_lwpr']) ## 특정 종목의 저가
# print(resp['output'][0]['stck_clpr']) ## 특정 종목의 종가
#print(resp['output'])



#pprint.pprint(resp)