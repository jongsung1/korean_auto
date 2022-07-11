from numpy import diff
import pandas as pd
import mojito
from KoreaStockAutoTrade import PERCENT, SYMBOL_LIST
import all_companys_list as COM ####### 종목 코드 리스트
import requests
import json
import yaml
import symbol_list as SL
from tqdm import tqdm ####### 진행률 표시 - for문에 사용
import openpyxl ####### 엑셀 write에 사용


with open('config.yaml', encoding='UTF-8') as f:
    _cfg = yaml.load(f, Loader=yaml.FullLoader)
APP_KEY = _cfg['APP_KEY']
APP_SECRET = _cfg['APP_SECRET']
ACCESS_TOKEN = ""
CANO = _cfg['CANO']
ACNT_PRDT_CD = _cfg['ACNT_PRDT_CD']
DISCORD_WEBHOOK_URL = _cfg['DISCORD_WEBHOOK_URL']
URL_BASE = _cfg['URL_BASE']
PERCENT = _cfg['PERCENT']   ##현재가가 최고가의 $percent% 범위 이내(0 ~ 1)
SYMBOL_LIST = _cfg['SYMBLO_LIST']   # 매수 희망 종목 리스트
K = _cfg['K']   ## 변동성 상수

def get_access_token():
    """토큰 발급"""
    headers = {"content-type":"application/json"}
    body = {"grant_type":"client_credentials",
    "appkey":APP_KEY, 
    "appsecret":APP_SECRET}
    PATH = "oauth2/tokenP"
    URL = f"{URL_BASE}/{PATH}"
    res = requests.post(URL, headers=headers, data=json.dumps(body))
    ACCESS_TOKEN = res.json()["access_token"]
    return ACCESS_TOKEN
    
def get_current_price(code="005930",symbol="aspr_unit"):
    """호가 aspr_unit
        BPS bps
        EPS eps
        PER per
        PBR pbr
        PRICE stck_prpr
        MARKET_CAP hts_avls
    """
    PATH = "/uapi/domestic-stock/v1/quotations/inquire-price"
    URL = f"{URL_BASE}/{PATH}"
    headers = {"Content-Type":"application/json", 
            "authorization": f"Bearer {ACCESS_TOKEN}",
            "appkey":APP_KEY,
            "appsecret":APP_SECRET,
            "tr_id":"FHKST01010100"}
    params = {
    "fid_cond_mrkt_div_code":"J",
    "fid_input_iscd":code,
    }
    res = requests.get(URL, headers=headers, params=params)
    #return int(res.json()['output'][symbol])
    try:
        return res.json()['output'][symbol]
    except KeyError:
        return 0

ACCESS_TOKEN = get_access_token()

####### 엑셀 저장을 위한 파트
row=1
wb = openpyxl.Workbook()
ws = wb.active

####### 엑셀 저장 함수
def write_excel(company_name,MARKET_CAP,PRICE,PER,PBR,EPS,BPS,row):
    ws.cell(row=row+1, column=1).value = company_name               ## 회사명
    ws.cell(row=row+1, column=2).value = MARKET_CAP                           ## 시가총액
    ws.cell(row=row+1, column=3).value = PRICE                       ## 현재가
    ws.cell(row=row+1, column=4).value = PER                           ## PER
    ws.cell(row=row+1, column=5).value = PBR                           ## PBR
    ws.cell(row=row+1, column=6).value = EPS                           ## EPS
    ws.cell(row=row+1, column=7).value = BPS                           ## BPS
    
####### 엑셀 헤더
write_excel("회사명","시가총액","현재가","PER","PBR","EPS","BPS",0)

# UNIT_PRICE=get_current_price("000660","aspr_unit")  ## SK하이닉스 종목의 호가
# EPS=get_current_price("000660","eps")   ## SK하이닉스 종목의 EPS
# BPS=get_current_price("000660","bps")   ## SK하이닉스 종목의 BPS
# PER=get_current_price("000660","per")   ## SK하이닉스 종목의 EPS
# PBR=get_current_price("000660","pbr")   ## SK하이닉스 종목의 PBR
# MARKET_CAP=get_current_price("000660","hts_avls")   ## SK하이닉스 종목의 시가총액
# PRICE=get_current_price("000660","stck_prpr")   ## SK하이닉스 종목의 현재가
# print("현재가 : %s , 시가총액 : %s , PER : %s , PBR : %s , EPS : %s , BPS : %s" % (PRICE,MARKET_CAP,PER,PBR,EPS,BPS))

### 저평가 종목 검색
for company in tqdm(COM.tickers.keys()):
    UNIT_PRICE=float(get_current_price(company,"aspr_unit"))   ## 호가
    EPS_TEMP=get_current_price(company,"eps")
    BPS_TEMP=get_current_price(company,"bps")
    PER_TEMP=get_current_price(company,"per")   
    PBR_TEMP=get_current_price(company,"pbr")
    MARKET_CAP_TEMP=get_current_price(company,"hts_avls")
    PRICE_TEMP=get_current_price(company,"stck_prpr")
    
    company_name=COM.tickers[company]   ## 회사명
    BPS=float(BPS_TEMP)   ## 종목의 BPS
    PER=float(PER_TEMP) ## 종목의 EPS
    EPS=float(EPS_TEMP)   ## 종목의 EPS
    PBR=float(PBR_TEMP)   ## 종목의 PBR
    PRICE=int(PRICE_TEMP)   ## 종목의 현재가
    MARKET_CAP=int(MARKET_CAP_TEMP)   ## 종목의 시가총액

    VAL_CHECK=EPS*BPS*PER*PBR*PRICE
    if VAL_CHECK != 0:
        #print(company_name)
        if MARKET_CAP > 2000:
            if PER < 10:
                if PBR <= 1:
#                    if EPS > PRICE:
                    ####### 엑셀 입력 함수 호출
                    write_excel(company_name,MARKET_CAP,PRICE,PER,PBR,EPS,BPS,row)
                    row = row + 1

wb.save(f"./financial_statements.xlsx")
